# run example:
# python import_commentary_edits_xlsx.py --commentary-db "..\commentary.db" --xlsx "..\08_export_commentary\commentary_draft.xlsx" --reset

# -*- coding: utf-8 -*-
#!/usr/bin/env python3
"""
import_commentary_edits_xlsx.py

Import human edits from an Excel (.xlsx) draft into commentary.db.

Edits table variants exist across iterations. This importer:
- Creates commentary_edits if missing (with created_at + updated_at defaults)
- If commentary_edits already exists and includes NOT NULL updated_at, it will populate it.

Requires:
  pip install openpyxl
"""

import argparse
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


def norm(v) -> str:
    return ("" if v is None else str(v)).strip()


def table_exists(con: sqlite3.Connection, table: str) -> bool:
    row = con.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?;",
        (table,),
    ).fetchone()
    return row is not None


def table_info(con: sqlite3.Connection, table: str) -> list[tuple]:
    # cid, name, type, notnull, dflt_value, pk
    return con.execute(f"PRAGMA table_info({table});").fetchall()


def ensure_schema(con: sqlite3.Connection):
    """
    Create commentary_edits if it doesn't exist.
    If it exists, we do NOT modify it; we just adapt inserts to its columns.
    """
    if table_exists(con, "commentary_edits"):
        return

    con.execute(
        """
        CREATE TABLE commentary_edits (
          id             INTEGER PRIMARY KEY AUTOINCREMENT,
          created_at     TEXT NOT NULL DEFAULT (datetime('now')),
          updated_at     TEXT NOT NULL DEFAULT (datetime('now')),

          model          TEXT NOT NULL,
          prompt_version TEXT NOT NULL,
          book_osis      TEXT NOT NULL,
          chapter        INTEGER NOT NULL,
          verse_start    INTEGER NOT NULL,
          verse_end      INTEGER NOT NULL,

          keep_drop      TEXT NOT NULL DEFAULT '',
          my_edit_en     TEXT NOT NULL DEFAULT '',
          my_edit_zh     TEXT NOT NULL DEFAULT '',
          notes          TEXT NOT NULL DEFAULT '',

          UNIQUE(model, prompt_version, book_osis, chapter, verse_start, verse_end)
        );
        """
    )
    con.commit()


def main() -> int:
    ap = argparse.ArgumentParser(description="Import edits from .xlsx into commentary.db")
    ap.add_argument("--commentary-db", required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", default="draft", help="Sheet name (default: draft)")
    ap.add_argument("--reset", action="store_true", help="Delete existing edits first")
    ap.add_argument("--verbose-bad-rows", action="store_true")
    args = ap.parse_args()

    db_path = Path(args.commentary_db).resolve()
    xlsx_path = Path(args.xlsx).resolve()

    con = sqlite3.connect(str(db_path))
    con.row_factory = sqlite3.Row
    try:
        ensure_schema(con)

        # detect commentary_edits columns (important for updated_at NOT NULL variants)
        info = table_info(con, "commentary_edits")
        cols = {r[1] for r in info}
        notnull = {r[1] for r in info if int(r[3] or 0) == 1}
        defaults = {r[1]: r[4] for r in info}  # may be None

        has_created_at = "created_at" in cols
        has_updated_at = "updated_at" in cols

        if args.reset:
            con.execute("DELETE FROM commentary_edits;")
            con.commit()

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        if args.sheet not in wb.sheetnames:
            raise SystemExit(f"ERROR: sheet '{args.sheet}' not found. Sheets: {wb.sheetnames}")
        ws = wb[args.sheet]

        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            raise SystemExit("ERROR: empty sheet")

        col_index = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

        required_cols = ["model", "prompt_version", "book_osis", "chapter", "verse_start", "verse_end"]
        for rc in required_cols:
            if rc not in col_index:
                raise SystemExit(f"ERROR: missing required column '{rc}' in xlsx header")

        keep_col = col_index.get("keep_drop")
        en_col = col_index.get("my_edit_en")
        zh_col = col_index.get("my_edit_zh")
        notes_col = col_index.get("notes")

        imported = 0
        skipped_blank = 0
        skipped_bad = 0

        # Build INSERT dynamically to satisfy schema differences (esp. updated_at NOT NULL)
        insert_cols = [
            "model",
            "prompt_version",
            "book_osis",
            "chapter",
            "verse_start",
            "verse_end",
            "keep_drop",
            "my_edit_en",
            "my_edit_zh",
            "notes",
        ]

        # If table has created_at/updated_at and they're NOT NULL (or present), we populate them.
        # (Even if they have defaults, setting them explicitly is fine.)
        # Use SQL datetime('now') so we don't worry about formatting.
        extra_sql_cols = []          # columns appended without params
        extra_sql_values = []        # their SQL expressions

        if has_created_at and ("created_at" in notnull or defaults.get("created_at") is None):
            extra_sql_cols.append("created_at")
            extra_sql_values.append("datetime('now')")

        if has_updated_at and ("updated_at" in notnull or defaults.get("updated_at") is None):
            extra_sql_cols.append("updated_at")
            extra_sql_values.append("datetime('now')")

        all_insert_cols = insert_cols + extra_sql_cols

        # VALUES placeholders for param-based cols, plus SQL expressions for timestamps
        param_placeholders = ",".join(["?"] * len(insert_cols))
        sql_expr_placeholders = ("," + ",".join(extra_sql_values)) if extra_sql_values else ""
        values_clause = f"{param_placeholders}{sql_expr_placeholders}"

        conflict_key = "model, prompt_version, book_osis, chapter, verse_start, verse_end"
        set_updates = [
            "keep_drop=excluded.keep_drop",
            "my_edit_en=excluded.my_edit_en",
            "my_edit_zh=excluded.my_edit_zh",
            "notes=excluded.notes",
        ]
        if has_updated_at:
            set_updates.append("updated_at=datetime('now')")

        insert_sql = f"""
        INSERT INTO commentary_edits ({", ".join(all_insert_cols)})
        VALUES ({values_clause})
        ON CONFLICT({conflict_key})
        DO UPDATE SET {", ".join(set_updates)}
        """

        for line_no, row in enumerate(it, start=2):
            def get(col):
                i = col_index.get(col)
                return row[i] if i is not None and i < len(row) else None

            model = norm(get("model"))
            prompt_version = norm(get("prompt_version"))
            book_osis = norm(get("book_osis"))

            try:
                chapter = int(norm(get("chapter")))
                verse_start = int(norm(get("verse_start")))
                verse_end = int(norm(get("verse_end")))
            except Exception:
                skipped_bad += 1
                if args.verbose_bad_rows:
                    print(f"SKIP line {line_no}: bad numeric key fields -> {row}")
                continue

            keep_drop = norm(row[keep_col]) if keep_col is not None else ""
            my_edit_en = norm(row[en_col]) if en_col is not None else ""
            my_edit_zh = norm(row[zh_col]) if zh_col is not None else ""
            notes = norm(row[notes_col]) if notes_col is not None else ""

            if not (keep_drop or my_edit_en or my_edit_zh or notes):
                skipped_blank += 1
                continue

            if not (model and prompt_version and book_osis):
                skipped_bad += 1
                if args.verbose_bad_rows:
                    print(f"SKIP line {line_no}: missing model/prompt_version/book_osis")
                continue

            params = [
                model,
                prompt_version,
                book_osis,
                chapter,
                verse_start,
                verse_end,
                keep_drop,
                my_edit_en,
                my_edit_zh,
                notes,
            ]
            con.execute(insert_sql, params)
            imported += 1

        con.commit()
        print(f"OK: imported edits: {imported}, skipped blank-edit rows: {skipped_blank}, skipped bad rows: {skipped_bad}")
        print(f"DB: {db_path}")
        return 0

    finally:
        con.close()


if __name__ == "__main__":
    raise SystemExit(main())
